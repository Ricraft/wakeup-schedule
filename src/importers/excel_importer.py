"""
Excel 导入器

从 Excel 课表文件导入课程数据
支持多种格式：
1. 强智系统导出的 Excel 格式
2. 标准课表格式（自动检测表头）
3. 简化格式（课程名 教师 地点）
"""

import uuid
import re
from typing import List, Tuple, Optional, Dict
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

try:
    from .base_importer import BaseImporter
    from .qiangzhi_importer import QiangZhiImporter
    from ..models.course_base import CourseBase
    from ..models.course_detail import CourseDetail
    from ..models.week_type import WeekType
    from ..utils.color_manager import ColorManager
except ImportError:
    from importers.base_importer import BaseImporter
    from importers.qiangzhi_importer import QiangZhiImporter
    from models.course_base import CourseBase
    from models.course_detail import CourseDetail
    from models.week_type import WeekType
    from utils.color_manager import ColorManager


class ExcelImporter(BaseImporter):
    """
    Excel 导入器
    
    智能解析 Excel 课表文件
    - 自动检测表头位置和格式
    - 支持强智系统导出格式
    - 支持标准课表格式
    """
    
    def __init__(self):
        """初始化 Excel 导入器"""
        self.color_manager = ColorManager()
        self.qiangzhi_importer = None  # 延迟初始化
        
        # 周次模式：{第2-16周 或 {第2-16周(单) 或 {第2-16周(双)
        self.week_pattern = re.compile(r'\{第(\d{1,2})[-]*(\d*)周(?:\((单|双)\))?')
        
        # 用于模糊匹配表头的关键字
        self.weekday_map = {
            "一": 1, "二": 2, "三": 3, "四": 4, "五": 5, "六": 6, "日": 7, "天": 7
        }
        
        # 星期映射（完整版）
        self.weekday_names = {
            "周一": 1, "星期一": 1, "Monday": 1, "Mon": 1,
            "周二": 2, "星期二": 2, "Tuesday": 2, "Tue": 2,
            "周三": 3, "星期三": 3, "Wednesday": 3, "Wed": 3,
            "周四": 4, "星期四": 4, "Thursday": 4, "Thu": 4,
            "周五": 5, "星期五": 5, "Friday": 5, "Fri": 5,
            "周六": 6, "星期六": 6, "Saturday": 6, "Sat": 6,
            "周日": 7, "星期日": 7, "Sunday": 7, "Sun": 7,
        }
    
    def get_supported_formats(self) -> List[str]:
        """获取支持的文件格式"""
        return ['.xlsx', '.xls']
    
    def validate(self, content: str) -> Tuple[bool, str]:
        """
        验证 Excel 文件路径是否有效
        
        Args:
            content: Excel 文件路径
            
        Returns:
            (是否有效, 错误消息)
        """
        if not content or not content.strip():
            return False, "文件路径为空"
        
        try:
            # 尝试打开文件
            workbook = openpyxl.load_workbook(content, read_only=True)
            
            # 检查是否有工作表
            if not workbook.sheetnames:
                return False, "Excel 文件中没有工作表"
            
            workbook.close()
            return True, ""
        except FileNotFoundError:
            return False, f"文件不存在: {content}"
        except Exception as e:
            return False, f"无法打开 Excel 文件: {str(e)}"
    
    def parse(self, file_path: str) -> Tuple[List[CourseBase], List[CourseDetail]]:
        """
        解析 Excel 文件
        
        Args:
            file_path: Excel 文件路径
            
        Returns:
            (CourseBase列表, CourseDetail列表)
            
        Raises:
            ValueError: 解析失败时抛出
        """
        # 验证文件
        valid, msg = self.validate(file_path)
        if not valid:
            raise ValueError(msg)
        
        # 打开 Excel 文件
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            
            # 检测格式类型
            format_type = self._detect_format(sheet)
            
            # 根据格式类型选择解析方法
            if format_type == "qiangzhi":
                import_beans = self._parse_qiangzhi_format(sheet)
            else:
                import_beans = self._parse_standard_format(sheet)
            
            # 转换为 CourseBase 和 CourseDetail
            course_bases, course_details = self._convert_to_courses(import_beans)
            
            workbook.close()
            return course_bases, course_details
        except Exception as e:
            raise ValueError(f"解析 Excel 文件失败: {str(e)}")
    
    def _detect_format(self, sheet: Worksheet) -> str:
        """
        检测 Excel 格式类型
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            格式类型：'qiangzhi' 或 'standard'
        """
        # 检查前几行是否包含强智系统特征
        rows = list(sheet.iter_rows(values_only=True, max_row=10))
        
        for row in rows:
            for cell in row:
                if cell and isinstance(cell, str):
                    # 检查是否包含强智系统特征
                    if '{第' in cell and '周' in cell:
                        return "qiangzhi"
        
        return "standard"
    
    def _parse_qiangzhi_format(self, sheet: Worksheet) -> List[dict]:
        """
        解析强智系统格式的 Excel - 增强版：支持自动定位和多行内容拆分
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            课程信息字典列表
        """
        import_beans = []
        
        # 获取所有行
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return import_beans
        
        # 智能查找表头行和列映射
        header_row_idx, weekday_col_map = self._find_header_and_columns(sheet)
        
        # 查找节次列（第一列通常是节次）
        section_col_idx = 0
        
        # 遍历数据行
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            
            # 提取节次信息
            current_section = self._extract_section_num(row[section_col_idx])
            if current_section is None:
                current_section = row_idx - header_row_idx
            
            # 遍历课程列
            for col_idx, cell_value in enumerate(row):
                # 跳过节次列
                if col_idx == section_col_idx:
                    continue
                
                # 跳过空单元格
                if not cell_value or str(cell_value).strip() == "":
                    continue
                
                # 获取星期
                day_of_week = weekday_col_map.get(col_idx, col_idx)
                
                # 解析单元格内容
                cell_text = str(cell_value)
                
                # 处理强智 Excel 格子内常见的 \n 分隔多门课
                if '{' in cell_text:
                    # 将内容按换行符拆分，每行可能是一门课
                    for sub_content in cell_text.split('\n'):
                        if '{' in sub_content:
                            # 转换为空格分隔
                            sub_content = sub_content.replace('\n', ' ').strip()
                            beans = self._parse_qiangzhi_cell(sub_content, current_section, day_of_week)
                            import_beans.extend(beans)
                else:
                    # 普通格式
                    cell_text = cell_text.replace('\n', ' ').strip()
                    beans = self._parse_qiangzhi_cell(cell_text, current_section, day_of_week)
                    import_beans.extend(beans)
        
        return import_beans
    
    def _find_header_and_columns(self, sheet: Worksheet) -> Tuple[int, Dict[int, int]]:
        """
        自动搜索包含"星期"或"周"的行，并识别各列对应的星期
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            (表头行索引, 列索引->星期数字的映射)
        """
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True, max_row=10)):
            col_map = {}
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                cell_str = str(cell)
                # 使用模糊匹配识别星期列
                for key, val in self.weekday_map.items():
                    if key in cell_str and ("周" in cell_str or "星期" in cell_str):
                        col_map[col_idx] = val
            
            if col_map:
                return row_idx, col_map
        
        # 默认回退：假设第1列=周一，第2列=周二等
        return 0, {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7}
    
    def _extract_section_num(self, text) -> Optional[int]:
        """
        从文本中提取节次编号
        
        Args:
            text: 包含节次信息的文本
            
        Returns:
            节次编号，如果无法提取则返回 None
        """
        if not text:
            return None
        match = re.search(r'(\d+)', str(text))
        return int(match.group(1)) if match else None
    
    def _parse_qiangzhi_cell(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析强智系统格式的单元格
        
        格式示例：
        "高等数学 {第1-16周 张老师 A101"
        "线性代数 {第1-16周(单) 李老师 B202"
        
        Args:
            content: 单元格文本
            section: 当前节次
            day: 星期几（1-7）
            
        Returns:
            课程信息字典列表
        """
        courses = []
        
        # 按空格分割
        parts = content.split()
        
        # 查找所有包含 { 的部分（周次信息）
        week_indices = []
        for i, part in enumerate(parts):
            if '{' in part:
                week_indices.append(i)
        
        # 如果没有找到周次信息，返回空列表
        if not week_indices:
            return courses
        
        # 解析每个课程
        for idx, week_idx in enumerate(week_indices):
            # 课程名称在周次信息之前
            if week_idx == 0:
                continue
            
            name = parts[week_idx - 1]
            time_info = parts[week_idx]
            
            # 教师和地点在周次信息之后
            teacher = ""
            location = ""
            
            # 确定这个课程的结束位置
            if idx < len(week_indices) - 1:
                next_week_idx = week_indices[idx + 1]
                end_idx = next_week_idx - 1
            else:
                end_idx = len(parts)
            
            # 提取教师和地点
            remaining_parts = parts[week_idx + 1:end_idx]
            if len(remaining_parts) >= 1:
                teacher = remaining_parts[0]
            if len(remaining_parts) >= 2:
                location = remaining_parts[1]
            
            courses.append({
                'name': name,
                'time_info': time_info,
                'teacher': teacher,
                'location': location,
                'section': section,
                'day': day
            })
        
        return courses
    
    def _parse_standard_format(self, sheet: Worksheet) -> List[dict]:
        """
        解析标准格式的 Excel（简化格式）
        
        格式：课程名 教师 地点
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            课程信息字典列表
        """
        import_beans = []
        
        # 获取所有行
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return import_beans
        
        # 查找表头行
        header_row_idx = 0
        weekday_col_map = {}
        
        for idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    for weekday_name, weekday_num in self.weekday_names.items():
                        if weekday_name in cell:
                            header_row_idx = idx
                            weekday_col_map[col_idx] = weekday_num
            
            if weekday_col_map:
                break
        
        # 如果没有找到表头，假设第一列是周一
        if not weekday_col_map:
            for i in range(1, 8):
                weekday_col_map[i] = i
        
        # 遍历数据行
        for row_idx in range(header_row_idx + 1, len(rows)):
            row = rows[row_idx]
            
            # 当前节次
            current_section = row_idx - header_row_idx
            
            # 遍历课程列
            for col_idx, cell_value in enumerate(row):
                # 跳过第一列（可能是节次列）
                if col_idx == 0:
                    continue
                
                # 跳过空单元格
                if not cell_value or str(cell_value).strip() == "":
                    continue
                
                # 获取星期
                day_of_week = weekday_col_map.get(col_idx, col_idx)
                
                # 解析单元格内容
                cell_text = str(cell_value).replace('\n', ' ').strip()
                beans = self._parse_simple_cell(cell_text, current_section, day_of_week)
                import_beans.extend(beans)
        
        return import_beans
    
    def _parse_simple_cell(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析简单格式的单元格
        
        格式：课程名 教师 地点
        
        Args:
            content: 单元格文本
            section: 当前节次
            day: 星期几（1-7）
            
        Returns:
            课程信息字典列表
        """
        parts = content.split()
        
        if not parts:
            return []
        
        name = parts[0]
        teacher = parts[1] if len(parts) >= 2 else ""
        location = parts[2] if len(parts) >= 3 else ""
        
        return [{
            'name': name,
            'time_info': '{第1-16周',  # 默认周次
            'teacher': teacher,
            'location': location,
            'section': section,
            'day': day
        }]
    
    def _parse_sheet(self, sheet: Worksheet) -> List[dict]:
        """
        解析工作表
        
        预设模板：
        - 第一行：标题行（可能包含"周一"、"周二"等）
        - 第一列：节次信息（可能包含"第1-2节"等）
        - 其他单元格：课程信息
        
        Args:
            sheet: openpyxl 工作表对象
            
        Returns:
            课程信息字典列表
        """
        import_beans = []
        
        # 获取所有行
        rows = list(sheet.iter_rows(values_only=True))
        
        if not rows:
            return import_beans
        
        # 遍历所有单元格
        for row_idx, row in enumerate(rows):
            # 跳过第一行（标题行）
            if row_idx == 0:
                continue
            
            # 当前行的节次（从第1节开始）
            current_section = row_idx
            
            for col_idx, cell_value in enumerate(row):
                # 跳过第一列（节次列）
                if col_idx == 0:
                    continue
                
                # 跳过空单元格
                if not cell_value or str(cell_value).strip() == "":
                    continue
                
                # 将单元格内容转换为空格分隔的字符串
                cell_text = str(cell_value).replace('\n', ' ').strip()
                
                # 推断星期（第2列=周一，第3列=周二，...）
                day_of_week = col_idx
                
                # 解析单元格内容
                # 复用 HTML 导入器的解析逻辑
                beans = self._parse_cell_content(cell_text, current_section, day_of_week)
                import_beans.extend(beans)
        
        return import_beans
    
    def _parse_cell_content(self, content: str, section: int, day: int) -> List[dict]:
        """
        解析单元格内容
        
        单元格格式示例：
        "高等数学 {第1-16周 张老师 A101"
        "线性代数 {第1-16周(单) 李老师 B202"
        
        Args:
            content: 单元格文本
            section: 当前节次
            day: 星期几（1-7）
            
        Returns:
            课程信息字典列表
        """
        courses = []
        
        # 按空格分割
        parts = content.split()
        
        # 查找所有包含 { 的部分（周次信息）
        week_indices = []
        for i, part in enumerate(parts):
            if '{' in part:
                week_indices.append(i)
        
        # 如果没有找到周次信息，尝试简单格式
        if not week_indices:
            # 简单格式：课程名 教师 地点
            if len(parts) >= 1:
                name = parts[0]
                teacher = parts[1] if len(parts) >= 2 else ""
                location = parts[2] if len(parts) >= 3 else ""
                
                courses.append({
                    'name': name,
                    'time_info': f'周{self._get_day_name(day)}{{第1-16周',  # 默认周次
                    'teacher': teacher,
                    'location': location,
                    'section': section,
                    'day': day
                })
            return courses
        
        # 解析每个课程
        for idx, week_idx in enumerate(week_indices):
            # 课程名称在周次信息之前
            if week_idx == 0:
                continue
            
            name = parts[week_idx - 1]
            time_info = parts[week_idx]
            
            # 教师和地点在周次信息之后
            teacher = None
            location = None
            
            # 确定这个课程的结束位置
            if idx < len(week_indices) - 1:
                next_week_idx = week_indices[idx + 1]
                end_idx = next_week_idx - 1
            else:
                end_idx = len(parts)
            
            # 提取教师和地点
            remaining_parts = parts[week_idx + 1:end_idx]
            if len(remaining_parts) >= 1:
                teacher = remaining_parts[0]
            if len(remaining_parts) >= 2:
                location = remaining_parts[1]
            
            courses.append({
                'name': name,
                'time_info': time_info,
                'teacher': teacher or "",
                'location': location or "",
                'section': section,
                'day': day
            })
        
        return courses
    
    def _convert_to_courses(self, import_beans: List[dict]) -> Tuple[List[CourseBase], List[CourseDetail]]:
        """
        将导入的数据转换为 CourseBase 和 CourseDetail
        
        Args:
            import_beans: 导入的课程信息列表
            
        Returns:
            (CourseBase列表, CourseDetail列表)
        """
        course_bases = []
        course_details = []
        course_id_map = {}  # 课程名称 -> 课程ID 的映射
        
        for bean in import_beans:
            name = bean['name']
            
            # 检查课程是否已存在
            if name not in course_id_map:
                # 创建新的 CourseBase
                course_id = str(uuid.uuid4())
                color = self.color_manager.get_color_for_course(name)
                course_base = CourseBase(
                    name=name,
                    color=color,
                    course_id=course_id
                )
                course_bases.append(course_base)
                course_id_map[name] = course_id
            else:
                course_id = course_id_map[name]
            
            # 解析时间信息
            time_info = self._parse_time_info(bean['time_info'])
            
            # 使用 Excel 中推断的星期（如果时间信息中没有）
            if time_info['day'] == 0:
                time_info['day'] = bean.get('day', 0)
            
            # 创建 CourseDetail
            course_detail = CourseDetail(
                course_id=course_id,
                teacher=bean['teacher'],
                location=bean['location'],
                day_of_week=time_info['day'],
                start_section=bean['section'],
                step=time_info['step'],
                start_week=time_info['start_week'],
                end_week=time_info['end_week'],
                week_type=time_info['week_type']
            )
            course_details.append(course_detail)
        
        return course_bases, course_details
    
    def _parse_time_info(self, time_info: str) -> dict:
        """
        解析时间信息
        
        Args:
            time_info: 时间信息字符串，如 "{第1-16周" 或 "{第1-16周(单)"
            
        Returns:
            包含 day, step, start_week, end_week, week_type 的字典
        """
        result = {
            'day': 0,
            'step': 1,
            'start_week': 1,
            'end_week': 20,
            'week_type': WeekType.EVERY_WEEK
        }
        
        # 解析周次
        match = self.week_pattern.search(time_info)
        if match:
            start_week_str = match.group(1)
            end_week_str = match.group(2)
            week_type_str = match.group(3)
            
            result['start_week'] = int(start_week_str)
            
            if end_week_str:
                result['end_week'] = int(end_week_str)
            else:
                result['end_week'] = result['start_week']
            
            # 解析单双周
            if week_type_str == '单':
                result['week_type'] = WeekType.ODD_WEEK
            elif week_type_str == '双':
                result['week_type'] = WeekType.EVEN_WEEK
        
        return result
    
    def _get_day_name(self, day: int) -> str:
        """
        将数字转换为中文星期
        
        Args:
            day: 星期数字（1-7）
            
        Returns:
            中文星期，如 "一"
        """
        day_names = ["", "一", "二", "三", "四", "五", "六", "日"]
        if 1 <= day <= 7:
            return day_names[day]
        return ""
