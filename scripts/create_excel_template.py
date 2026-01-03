"""
创建标准 Excel 课表模板

运行此脚本会在 examples/ 目录下生成一个标准的 Excel 课表模板文件
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path


def create_template():
    """创建标准 Excel 课表模板"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课表"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 25
    
    # 设置行高
    ws.row_dimensions[1].height = 25
    for row in range(2, 14):
        ws.row_dimensions[row].height = 60
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_font = Font(name='微软雅黑', size=10)
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    section_font = Font(name='微软雅黑', size=11, bold=True)
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # 填充表头
    headers = ['节次', '周一', '周二', '周三', '周四', '周五', '周六', '周日']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 填充节次列
    for row_idx in range(2, 14):
        section = row_idx - 1
        cell = ws.cell(row=row_idx, column=1, value=section)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = section_alignment
        cell.border = border
    
    # 填充示例数据
    examples = [
        (2, 2, "高等数学\n{第1-16周\n张老师\nA101"),
        (2, 4, "线性代数\n{第1-16周(单)\n李老师\nB202"),
        (3, 3, "大学物理\n{第1-16周\n王老师\nC303"),
        (4, 2, "英语\n{第1-16周\n赵老师\nD404"),
        (5, 5, "计算机基础\n{第1-16周(双)\n刘老师\nE505"),
    ]
    
    for row, col, value in examples:
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = border
    
    # 填充空单元格（添加边框）
    for row in range(2, 14):
        for col in range(2, 9):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = ""
            cell.font = cell_font
            cell.alignment = cell_alignment
            cell.border = border
    
    # 创建说明工作表
    ws_info = wb.create_sheet("使用说明")
    ws_info.column_dimensions['A'].width = 80
    
    instructions = [
        "WakeUp 课程表 - Excel 导入模板使用说明",
        "",
        "一、填写格式",
        "",
        "1. 完整格式（推荐）：",
        "   课程名",
        "   {第1-16周",
        "   教师名",
        "   教室",
        "",
        "   示例：",
        "   高等数学",
        "   {第1-16周",
        "   张老师",
        "   A101",
        "",
        "2. 简化格式：",
        "   课程名 教师名 教室",
        "   （默认周次为 1-16 周）",
        "",
        "3. 单双周标记：",
        "   - 单周：{第1-16周(单)",
        "   - 双周：{第1-16周(双)",
        "   - 每周：{第1-16周",
        "",
        "二、注意事项",
        "",
        "1. 不要修改表头行（第一行）",
        "2. 不要修改节次列（第一列）",
        "3. 每个单元格只填写一门课程",
        "4. 避免使用合并单元格",
        "5. 如果一门课程有多个时间段，请分别填写在对应的单元格中",
        "",
        "三、导入步骤",
        "",
        "1. 按照格式填写课程信息",
        "2. 保存文件",
        '3. 在 WakeUp 中选择"文件" → "导入" → "从 Excel 导入"',
        "4. 选择保存的文件",
        "5. 系统会自动解析并导入课程",
        "",
        "四、常见问题",
        "",
        "Q: 为什么有些课程没有导入？",
        "A: 请检查单元格格式是否正确，特别是周次格式（{第1-16周）",
        "",
        "Q: 如何导入单双周课程？",
        "A: 在周次信息后添加 (单) 或 (双)，如：{第1-16周(单)",
        "",
        "Q: 可以修改模板的列数吗？",
        "A: 不建议修改。如果需要，请确保保留表头行的星期信息",
        "",
        "五、技术支持",
        "",
        "如果遇到问题，请查看 docs/IMPORT_GUIDE.md 获取详细帮助",
        "或访问项目 GitHub 仓库提交 Issue",
    ]
    
    for row_idx, text in enumerate(instructions, start=1):
        cell = ws_info.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(name='微软雅黑', size=14, bold=True, color='4472C4')
        elif text.startswith(('一、', '二、', '三、', '四、', '五、')):
            cell.font = Font(name='微软雅黑', size=12, bold=True)
        elif text.startswith(('Q:', 'A:')):
            cell.font = Font(name='微软雅黑', size=10, bold=True)
        else:
            cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 保存文件
    output_dir = Path(__file__).parent.parent / "examples"
    output_dir.mkdir(exist_ok=True)
    output_file = output_dir / "schedule_template.xlsx"
    
    wb.save(output_file)
    print(f"✓ Excel 模板已创建: {output_file}")
    print(f"  文件大小: {output_file.stat().st_size / 1024:.2f} KB")
    print(f"  工作表数量: {len(wb.sheetnames)}")
    print(f"  工作表名称: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    print("=" * 60)
    print("创建 WakeUp 课程表 Excel 模板")
    print("=" * 60)
    print()
    
    try:
        create_template()
        print()
        print("=" * 60)
        print("✓ 模板创建成功！")
        print("=" * 60)
    except Exception as e:
        print()
        print("=" * 60)
        print(f"✗ 创建失败: {e}")
        print("=" * 60)
        import traceback
        traceback.print_exc()
